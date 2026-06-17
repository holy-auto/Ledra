#!/usr/bin/env python3
"""
出願管理台帳 Excel(.xlsx) を生成。出願日を入力すると期限を自動計算する数式入り。

期限式（Excel/LibreOffice の EDATE を使用。要・正確な月数/日数を最新仕様で確認）:
  公開予定        = EDATE(出願日, 18)
  審査請求期限    = EDATE(出願日, 36)        # 出願日+3年
  30条証明書提出  = 出願日 + 30              # 主張時のみ。実務上30日(要確認)
  優先期限        = EDATE(出願日, 12)        # パリ優先・PCT/外国出願の最終
  PCT移行期限     = EDATE(出願日, 31)        # 優先日基準(国により30/31月)

使い方:  pip install openpyxl && python3 make_xlsx.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "出願管理台帳.xlsx")

HEAD = ["整理番号", "発明", "出願人", "代表者", "連絡先", "出願種別", "出願日",
        "公開予定", "審査請求期限", "30条主張", "30条証明書提出期限",
        "優先期限", "PCT移行期限", "登録日", "次回年金期限", "ステータス", "メモ"]
# 列記号
C = {h: get_column_letter(i + 1) for i, h in enumerate(HEAD)}

ROWS = [
    ["HA-ZKP-001", "発明06 ZKP選択的開示", "株式会社HOLY", "堀越 友輔", "info@ledra.co.jp",
     "国内", None, None, None, "不要", None, None, None, None, None, "出願準備", "本命・公開凍結維持"],
    ["HA-GATE-001", "発明07 確定ゲート", "株式会社HOLY", "堀越 友輔", "info@ledra.co.jp",
     "国内", None, None, None, "不要", None, None, None, None, None, "出願準備", "防御の堀・公開凍結維持"],
    ["HA-META-001", "発明01 メタアンカー", "株式会社HOLY", "堀越 友輔", "info@ledra.co.jp",
     "国内", None, None, None, "要(自社開示)", None, None, None, None, None, "出願準備(30条要確定)",
     "独立項は機構限定／外国は猶予なし=機構限定で出願"],
    ["HA-PIPE-001", "統合(任意)", "株式会社HOLY", "堀越 友輔", "info@ledra.co.jp",
     "国内", None, None, None, "要確認", None, None, None, None, None, "保留(単一性)", "個別3出願を優先"],
]

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "管理台帳"

    hfill = PatternFill("solid", fgColor="DDDDDD")
    for ci, h in enumerate(HEAD, 1):
        c = ws.cell(1, ci, h)
        c.font = Font(bold=True)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    date_fmt = "yyyy/mm/dd"
    for ri, row in enumerate(ROWS, start=2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val)
        g = f"{C['出願日']}{ri}"
        j = f"{C['30条主張']}{ri}"
        ws[f"{C['公開予定']}{ri}"] = f'=IF({g}="","",EDATE({g},18))'
        ws[f"{C['審査請求期限']}{ri}"] = f'=IF({g}="","",EDATE({g},36))'
        ws[f"{C['30条証明書提出期限']}{ri}"] = f'=IF({g}="","",IF(LEFT({j},1)="要",{g}+30,""))'
        ws[f"{C['優先期限']}{ri}"] = f'=IF({g}="","",EDATE({g},12))'
        ws[f"{C['PCT移行期限']}{ri}"] = f'=IF({g}="","",EDATE({g},31))'
        for col in ["出願日", "公開予定", "審査請求期限", "30条証明書提出期限",
                    "優先期限", "PCT移行期限", "登録日", "次回年金期限"]:
            ws[f"{C[col]}{ri}"].number_format = date_fmt

    widths = {"整理番号": 13, "発明": 20, "出願人": 12, "代表者": 10, "連絡先": 18,
              "出願種別": 9, "出願日": 12, "公開予定": 12, "審査請求期限": 13,
              "30条主張": 12, "30条証明書提出期限": 16, "優先期限": 12, "PCT移行期限": 13,
              "登録日": 12, "次回年金期限": 13, "ステータス": 16, "メモ": 34}
    for h, w in widths.items():
        ws.column_dimensions[C[h]].width = w
    ws.freeze_panes = "B2"

    # 説明シート
    ws2 = wb.create_sheet("説明")
    notes = [
        ["出願管理台帳 — 説明", ""],
        ["", ""],
        ["使い方", "各行の『出願日』に日付を入れると、公開予定/審査請求期限/30条/優先/PCT移行が自動計算される。"],
        ["公開予定", "= EDATE(出願日,18)  出願公開(出願日+1年6月で自動公開)"],
        ["審査請求期限", "= EDATE(出願日,36)  出願日+3年。徒過で取下げ擬制"],
        ["30条証明書提出期限", "= 出願日+30  30条主張時のみ(発明01)。実務上30日・要確認"],
        ["優先期限", "= EDATE(出願日,12)  パリ優先・PCT/外国出願の最終期限"],
        ["PCT移行期限", "= EDATE(出願日,31)  優先日基準(国により30/31月)"],
        ["年金期限", "登録後の特許料納付期限。登録日確定後に手動記入"],
        ["", ""],
        ["出願人", "株式会社HOLY（登記商号。願書の出願人名は登記と完全一致が必須）"],
        ["代表者", "堀越 友輔"],
        ["連絡先", "info@ledra.co.jp"],
        ["所在地", "〒107-0061 東京都港区北青山1-3-1 アールキューブ青山3F"],
        ["注意", "月数/日数・現行料金は特許庁(JPO)で要確認。30条は発明01のみ(自社開示)。"],
    ]
    for r, (a, b) in enumerate(notes, 1):
        ws2.cell(r, 1, a).font = Font(bold=(r == 1))
        ws2.cell(r, 2, b)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 80

    wb.save(OUT)
    print(f"wrote {os.path.relpath(OUT, HERE)}  (sheets: 管理台帳, 説明 / 出願日入力で期限自動計算)")

if __name__ == "__main__":
    main()
